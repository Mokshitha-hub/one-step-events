from flask import Flask, render_template, request, redirect, url_for, session
import pandas as pd
import os
from uuid import uuid4
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

app = Flask(__name__)
app.secret_key = "one_step_events_secret_key"

FILE_NAME = "event_enquiries.xlsx"

# =========================
# 🔐 ONLY ADMIN LOGIN
# =========================
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"


# =========================
# 🔐 LOGIN PAGE (HIDDEN)
# =========================
@app.route("/secure-login-x9k2")   # hidden login URL
def login_page():
    if "user" in session:
        return redirect(url_for("dashboard"))
    return render_template("login.html")


@app.route("/login", methods=["POST"])
def login():
    username = request.form["username"]
    password = request.form["password"]

    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        session["user"] = username
        return redirect(url_for("dashboard"))

    return "❌ Invalid Username or Password"


@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login_page"))


# =========================
# 📊 LOAD DATA
# =========================
def load_enquiries():
    schema_changed = False

    if os.path.exists(FILE_NAME):
        df = pd.read_excel(FILE_NAME)
    else:
        df = pd.DataFrame(columns=[
            "Enquiry ID",
            "Full Name",
            "Mobile Number",
            "Email",
            "Expected Date",
            "Event Type",
            "Budget",
            "Vision",
            "Status",
        ])

    if "Status" not in df.columns:
        df["Status"] = "Pending"
        schema_changed = True

    df["Status"] = df["Status"].fillna("Pending")

    if schema_changed:
        save_enquiries(df)

    return df


# =========================
# 💾 SAVE DATA
# =========================
def save_enquiries(df):
    df.to_excel(FILE_NAME, index=False)

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    if "EventEnquiries" in ws.tables:
        del ws.tables["EventEnquiries"]

    table_range = f"A1:I{ws.max_row}"
    tab = Table(displayName="EventEnquiries", ref=table_range)

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(FILE_NAME)


# =========================
# 🏠 HOME PAGE (NO DASHBOARD HERE)
# =========================
@app.route("/")
def home():
    return render_template("index.html")


# =========================
# 📊 DASHBOARD (PROTECTED)
# =========================
@app.route("/dashboard")
def dashboard():

    if "user" not in session:
        return redirect(url_for("login_page"))

    df = load_enquiries()
    enquiries = df.to_dict(orient="records")

    approved_count = int((df["Status"] == "Approved").sum())
    pending_count = int((df["Status"] != "Approved").sum())

    return render_template(
        "dashboard.html",
        enquiries=enquiries,
        approved_count=approved_count,
        pending_count=pending_count
    )


# =========================
# 📝 SUBMIT ENQUIRY
# =========================
@app.route("/submit-enquiry", methods=["POST"])
def submit_enquiry():

    full_name = request.form["full_name"]
    mobile = request.form["mobile"]
    email = request.form["email"]
    expected_date = request.form["expected_date"]
    event_type = request.form["event_type"]
    budget = request.form["budget"]
    vision = request.form["vision"]

    df = load_enquiries()

    new_row = pd.DataFrame([{
        "Enquiry ID": str(uuid4()),
        "Full Name": full_name,
        "Mobile Number": mobile,
        "Email": email,
        "Expected Date": expected_date,
        "Event Type": event_type,
        "Budget": budget,
        "Vision": vision,
        "Status": "Pending",
    }])

    df = pd.concat([df, new_row], ignore_index=True)
    save_enquiries(df)

    return render_template("success.html")


# =========================
# 🔄 UPDATE STATUS
# =========================
@app.route("/update-status", methods=["POST"])
def update_status():

    if "user" not in session:
        return redirect(url_for("login_page"))

    enquiry_id = request.form["enquiry_id"]
    new_status = request.form.get("status_action", "pending")

    if new_status == "approve":
        new_status = "Approved"
    else:
        new_status = "Pending"

    df = load_enquiries()
    df.loc[df["Enquiry ID"].astype(str) == str(enquiry_id), "Status"] = new_status
    save_enquiries(df)

    return redirect(url_for("dashboard"))


# =========================
# 🚀 RUN APP
# =========================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5002, debug=True)